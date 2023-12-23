from collections import Counter
from itertools import combinations
import openpyxl
from datetime import datetime
import random


def ler_arquivo_xlsx(nome_arquivo):
    workbook = openpyxl.load_workbook(nome_arquivo + ".xlsx")
    planilha = workbook.active

    # Listas para armazenar números e datas
    numeros = []
    datas = []

    # Iterar pelas linhas e pegar números das colunas específicas (C a H) e as datas da coluna B
    for i, row in enumerate(
        planilha.iter_rows(
            min_row=1, max_row=planilha.max_row, min_col=2, max_col=8, values_only=True
        )
    ):
        if i != 0:  # Ignorar a primeira linha com cabeçalhos
            datas.append(
                datetime.strptime(row[0], "%d/%m/%Y")
            )  # Convertendo a string para objeto datetime
            numeros.extend(row[1:])  # Colunas C a H contêm os números

    return numeros, datas


# Nome do arquivo Excel (sem a extensão)
nome_arquivo = "MegaSena"

# Ler todos os números e datas do arquivo Excel
numeros, datas = ler_arquivo_xlsx(nome_arquivo)

# Calcular os números mais sorteados
contagem_numeros = Counter(numeros)
numeros_mais_sorteados = contagem_numeros.most_common(60)

# Mostrar os números por ordem de mais sorteados e suas probabilidades
print("Números por ordem de mais sorteados:")
for numero, frequencia in numeros_mais_sorteados:
    probabilidade = frequencia / len(numeros)
    print(
        f"Número: {numero} - Frequência: {frequencia} - Probabilidade: {probabilidade:.6f}"
    )


# Calcular a probabilidade considerando o tempo desde o último sorteio
def calcular_probabilidade(numeros_mais_sorteados, datas):
    total = sum(frequencia for _, frequencia in numeros_mais_sorteados)

    # Calculando a última data de cada número
    ultima_data = {}
    for numero in set(numeros):
        ultima_data[numero] = max(
            (data for num, data in zip(numeros, datas) if num == numero), default=None
        )

    # Calculando a probabilidade considerando o tempo desde o último sorteio
    probabilidades = {}
    for numero, frequencia in numeros_mais_sorteados:
        dias_desde_ultimo_sorteio = (
            (datetime.now() - ultima_data[numero]).days
            if ultima_data[numero]
            else float("inf")
        )
        peso_temporal = 1 / (
            1 + dias_desde_ultimo_sorteio
        )  # Quanto maior o peso, mais tempo desde o último sorteio
        probabilidades[numero] = (frequencia / total) * peso_temporal

    return probabilidades


# Calcular a probabilidade considerando o tempo desde o último sorteio
probabilidades = calcular_probabilidade(numeros_mais_sorteados, datas)


# Função para gerar jogos únicos
def gerar_jogo_unico(probabilidades):
    jogo = random.choices(
        list(probabilidades.keys()), weights=list(probabilidades.values()), k=6
    )
    while len(set(jogo)) < 6:
        novo_numero = random.choice(list(probabilidades.keys()))
        if novo_numero not in jogo:
            jogo[random.randint(0, 5)] = novo_numero
    return jogo


# Gerar os três jogos mais prováveis sem repetição
jogos_mais_provaveis_sem_repeticao = []
while len(jogos_mais_provaveis_sem_repeticao) < 3:
    jogo = gerar_jogo_unico(probabilidades)
    if jogo not in jogos_mais_provaveis_sem_repeticao:
        jogos_mais_provaveis_sem_repeticao.append(jogo)

# Calcular a probabilidade de cada jogo sem repetição
probabilidade_jogos_sem_repeticao = []
for jogo in jogos_mais_provaveis_sem_repeticao:
    probabilidade_jogo = sum(probabilidades[numero] for numero in jogo) / 6
    probabilidade_jogos_sem_repeticao.append((jogo, probabilidade_jogo))

# Mostrar os três jogos mais prováveis sem repetição
print("\nTrês jogos mais prováveis para o próximo sorteio (sem repetição):")
for index, (jogo, probabilidade) in enumerate(
    probabilidade_jogos_sem_repeticao, start=1
):
    print(f"Jogo {index}: {jogo} - Probabilidade: {probabilidade:.6f}")

# Implementação para encontrar os 10 números mais e menos sorteados que não saíram há mais tempo
ultima_data = {}
for numero in set(numeros):
    ultima_data[numero] = max(
        (data for num, data in zip(numeros, datas) if num == numero), default=None
    )

numeros_ordenados_por_ultimo_sorteio = sorted(
    probabilidades.keys(),
    key=lambda num: (probabilidades[num], -1 * (datetime.now() - ultima_data[num]).days),
    reverse=True,
)

# Pegar os 10 números mais sorteados que não saíram por mais tempo
top_10_mais_sorteados_e_com_mais_tempo = [
    numero
    for numero in numeros_ordenados_por_ultimo_sorteio
    if numero not in [num for num, _ in numeros_mais_sorteados[:10]]
][:10]

# Pegar os 10 números menos sorteados que não saíram por mais tempo
top_10_menos_sorteados_e_com_mais_tempo = [
    numero
    for numero in reversed(numeros_ordenados_por_ultimo_sorteio)
    if numero not in [num for num, _ in numeros_mais_sorteados[-10:]]
][:10]

# Mostrar os números mais e menos sorteados que não saíram por mais tempo
print("\nDez números mais sorteados que não saíram por mais tempo:")
print(top_10_mais_sorteados_e_com_mais_tempo)

print("\nDez números menos sorteados que não saíram por mais tempo:")
print(top_10_menos_sorteados_e_com_mais_tempo)
